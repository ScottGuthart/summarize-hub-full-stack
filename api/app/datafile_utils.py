from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
import csv
import os
from pathlib import Path


def append_column_to_excel_preserve_format(
    file_path,
    column_name,
    column_data,
    output_path=None,
    sheet_name=None
):
    """
    Append a new column to an Excel file while preserving original formatting.
    
    Args:
        file_path (str): Path to the input Excel file
        column_name (str): Name of the new column to add
        column_data (list): Data for the new column
        output_path (str, optional): Path to save the modified file. 
            If None, overwrites the input file
        sheet_name (str, optional): Name of the sheet to modify. 
            If None, uses the active sheet
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Load the workbook and sheet
        wb = load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # Find the next empty column
        last_col = ws.max_column
        new_col_letter = get_column_letter(last_col + 1)
        
        # Add the header
        header_cell = ws[f"{new_col_letter}1"]
        header_cell.value = column_name
        
        # Copy header formatting if available
        if last_col > 0:
            prev_header = ws[f"{get_column_letter(last_col)}1"]
            if prev_header.has_style:
                header_cell._style = copy(prev_header._style)
        
        # Add the data
        for row_num, value in enumerate(column_data, start=2):
            new_cell = ws[f"{new_col_letter}{row_num}"]
            new_cell.value = value
            
            # Copy formatting from adjacent column if it exists
            if last_col > 0:
                prev_cell = ws[f"{get_column_letter(last_col)}{row_num}"]
                if prev_cell.has_style:
                    new_cell._style = copy(prev_cell._style)
        
        # Auto-adjust column width to fit content
        max_length = len(str(column_name))
        for cell in ws[new_col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[new_col_letter].width = max_length + 2
        
        # Save the workbook
        save_path = output_path if output_path else file_path
        wb.save(save_path)
        return True
        
    except Exception as e:
        print(f"Error appending column: {str(e)}")
        return False


def append_column_to_csv(
    file_path,
    column_name,
    column_data,
    output_path=None,
    delimiter=','
):
    """
    Append a new column to a CSV file.
    
    Args:
        file_path (str): Path to the input CSV file
        column_name (str): Name of the new column to add
        column_data (list): Data for the new column
        output_path (str, optional): Path to save the modified file.
            If None, overwrites the input file
        delimiter (str, optional): CSV delimiter character. Defaults to comma
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Read the existing CSV file
        rows = []
        with open(file_path, 'r', newline='') as csvfile:
            reader = csv.reader(csvfile, delimiter=delimiter)
            rows = list(reader)
        
        if not rows:
            # If file is empty, create with header
            rows = [[column_name]]
            rows.extend([[value] for value in column_data])
        else:
            # Add header to first row
            rows[0].append(column_name)
            
            # Add data to subsequent rows
            for i, value in enumerate(column_data, start=1):
                # Extend row if needed
                while len(rows) <= i:
                    rows.append([''] * (len(rows[0]) - 1))
                rows[i].append(str(value))
            
            # Pad any remaining rows with empty values
            for i in range(len(column_data) + 1, len(rows)):
                rows[i].append('')
        
        # Write to temporary file first
        temp_path = output_path if output_path else file_path + '.tmp'
        with open(temp_path, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile, delimiter=delimiter)
            writer.writerows(rows)
        
        # If successful and no output_path specified, replace original file
        if not output_path:
            os.replace(temp_path, file_path)
            
        return True
        
    except Exception as e:
        print(f"Error appending column: {str(e)}")
        if not output_path and os.path.exists(file_path + '.tmp'):
            os.remove(file_path + '.tmp')
        return False


def append_column(
    file_path,
    column_name,
    column_data,
    output_path=None,
    sheet_name=None,
    delimiter=','
):
    """
    Append a column to either an Excel or CSV file, automatically detecting the type.
    
    Args:
        file_path (str): Path to the input file (Excel or CSV)
        column_name (str): Name of the new column to add
        column_data (list): Data for the new column
        output_path (str, optional): Path to save the modified file.
            If None, overwrites the input file
        sheet_name (str, optional): Sheet name for Excel files.
            If None, uses the active sheet
        delimiter (str, optional): Delimiter for CSV files. Defaults to comma
    
    Returns:
        bool: True if successful, False otherwise
        
    Raises:
        ValueError: If the file type is not supported
    """
    file_ext = Path(file_path).suffix.lower()
    
    if file_ext in ['.xlsx', '.xls']:
        return append_column_to_excel_preserve_format(
            file_path=file_path,
            column_name=column_name,
            column_data=column_data,
            output_path=output_path,
            sheet_name=sheet_name
        )
    elif file_ext == '.csv':
        return append_column_to_csv(
            file_path=file_path,
            column_name=column_name,
            column_data=column_data,
            output_path=output_path,
            delimiter=delimiter
        )
    else:
        raise ValueError(
            f"Unsupported file type: {file_ext}. "
            "Supported types are: .xlsx, .xls, .csv"
        )


# Example usage:
"""
# Automatically handles both Excel and CSV files
file_path = 'data.xlsx'  # or 'data.csv'
new_column = 'New Column'
data = [1, 2, 3, 4, 5]  # Must match rows in file

# Basic usage
append_column(file_path, new_column, data)

# Advanced usage with all options
append_column(
    file_path=file_path,
    column_name=new_column,
    column_data=data,
    output_path='output.xlsx',  # or 'output.csv'
    sheet_name='Sheet1',  # for Excel only
    delimiter=';'  # for CSV only
)
""" 